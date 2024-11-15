from django.shortcuts import render, HttpResponse, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from .forms import RegisterForms
from django.http import FileResponse, StreamingHttpResponse
from django.contrib import admin
from django.contrib import messages
from .models import CustomUser
from django.core.cache import cache
from django.core.validators import validate_email 
from .models import InitialReport
from datetime import date,  timedelta
from django.db.models.functions import TruncMonth
from django.utils.timezone import now
from django.db.models import Count
from django.shortcuts import get_object_or_404, redirect
import random
import string
from django.core.mail import send_mail
from django.conf import settings
from django.http import JsonResponse
from django.contrib.auth.hashers import make_password
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.utils import timezone
from datetime import datetime
from datetime import  date
import json
import time
from django.core.exceptions import ValidationError
from django.utils.dateparse import parse_date
from django.core.validators import validate_integer
import time
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def export_initial_report(request):
    button_value = request.GET.get('value', None)

    print("Button value:", button_value)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Initial Report Data"

    headers = [
        "No.", "Station", "Exact Location/ Address of Fire Incident", "Team", "Time Reported", "Date Reported", "Type Of", "Name of Owner",
        "Alarm Status", "Alarm Declared By", "Time of Arrival At Scene", "Time of Fire Under Control", 
        "Date of Fire Under Control", "Fire Under Control Declared By", "Time of Fire Out", 
        "Date of Fire Out", "Fire Out Declared By", "Estimated Damages", "No. of Fatalities",
        "No. of Injured", "No. of Families Affected", "No. of Establishments", 
        "No. of Fire Trucks", "Ground Commander", "Commander Contact Number", 
        "Safety Officer", "Officer Contact Number", "Name of Sender", "Sender Contact Number", 
        "Remarks"
    ]

    fields = [
        'where', 'team', 'time_reported', 'date_reported', 'involved', 'name_of_owner',
        'alarm_status', 'alarm_declared_by', 'time_of_arrival', 'time_of_fire_under_control',
        'date_of_fire_under_control', 'fire_under_control_declared_by', 'time_of_fire_out',
        'date_of_fire_out', 'fire_out_declared_by', 'estimated_damages', 'no_of_fatality',
        'no_of_injured', 'no_of_families_affected', 'no_of_establishments', 'no_of_fire_trucks',
        'ground_commander', 'commander_contact_number', 'safety_officer', 'officer_contact_number',
        'name_of_sender', 'sender_contact_number', 'status'
    ]

    header_font = Font(bold=True, color="000000")

    data_fill = PatternFill(start_color="DF4B40", end_color="DF4B40", fill_type="solid")
 
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thick_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    for col_num, header_item in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header_item
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thick_border

    ws.row_dimensions[1].height = 30

    sy_text = ws.cell(row=2, column=1, value="S/Y 2024")
    sy_text.font = Font(bold=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sy_text.alignment = center_alignment  
    sy_text.border = thick_border

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = yellow_fill
        cell.alignment = center_alignment
        cell.border = thick_border

    reports = InitialReport.objects.filter(status='Case Closed').values(*fields)

    for row_num, report in enumerate(reports, 4):

        no_value = row_num - 3  
        
        station_value = "Santa Rosa City"
        
        no_cell = ws.cell(row=row_num, column=1, value=no_value)
        no_cell.fill = data_fill
        no_cell.alignment = center_alignment
        no_cell.border = thick_border  
        
        station_cell = ws.cell(row=row_num, column=2, value=station_value)  
        station_cell.fill = data_fill
        station_cell.alignment = center_alignment  
        station_cell.border = thick_border
        
        for col_num, field in enumerate(fields, 3):
            value = report[field]

            if isinstance(value, datetime):
                value = value.replace(tzinfo=None)  
            
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = data_fill
            cell.alignment = center_alignment
            cell.border = thick_border

    for col in range(2, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    ws.column_dimensions['A'].width = 5  

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="initial_report_data.xlsx"'
    wb.save(response)
    return response



def change_password(request):
    if request.method == 'POST':
        new_password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')

        if new_password != confirm_password:
            messages.error(request, 'Passwords do not match.')
            return redirect(request.META.get('HTTP_REFERER', 'analytics'))

        user = request.user

        user.set_password(new_password)
        user.save()

        update_session_auth_hash(request, user)

        send_mail(
            'Password Reset Successful',
            f'Hello {user.get_full_name() or user.name},\n\nYour password has been successfully changed.\n\nPlease keep it safe and secure.',
            settings.DEFAULT_FROM_EMAIL,
            [user.email],
            fail_silently=False,
        )
        
        messages.success(request, 'Password reset successful! A confirmation email has been sent.')

        return redirect(request.META.get('HTTP_REFERER', 'analytics'))

    return render(request, 'analytics.html')


def generate_random_password(length=10):
    """Generates a random password."""
    characters = string.ascii_letters + string.digits + string.punctuation
    return ''.join(random.choice(characters) for i in range(length))


def forgot_password_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        new_password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')

        if new_password != confirm_password:
            messages.error(request, 'Passwords do not match.')
            return redirect('forgot_password')

        try:
            user = CustomUser.objects.get(email=email)

            user.set_password(new_password)
            user.save()

            send_mail(
                'Password Reset Successful',
                f'Hello {user.name},\n\nYour password has been successfully changed.\n\nPlease keep it safe and secure.',
                settings.DEFAULT_FROM_EMAIL,
                [email],
                fail_silently=False,
            )

            messages.success(request, 'Password reset successful! A confirmation email has been sent.')
            return redirect('login')

        except CustomUser.DoesNotExist:
            messages.error(request, 'Email address not found.')
            return redirect('forgot_password')

    return render(request, 'forgot_password.html')


@login_required
def register_view(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        role = request.POST.get('role')
        contact_number = request.POST.get('contact_number')  

        if not all([name, email, role, contact_number]):
            messages.error(request, 'All fields are required.')
        else:
            try:
                # Validate the contact number format (optional)
                if not contact_number.isdigit() or len(contact_number) not in [10, 11]:
                    raise ValidationError("Contact number must be 10-11 digits.")

                generated_password = generate_random_password()

                new_user = CustomUser.objects.create_user(
                    name=name,
                    email=email,
                    role=role,
                    contact_number=contact_number,
                    password=generated_password
                )
                
                send_mail(
                    'Your Account Password',
                    f'Hello {name},\n\nYour account has been created. Your password is: {generated_password}\nYou can change your password after logging in.',
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=False,
                )

                messages.success(request, 'Account has been added successfully, and the password has been sent to the user\'s email.')
            except ValidationError as e:
                messages.error(request, f'Validation error: {e.message}')
            except Exception as e:
                messages.error(request, f'Error adding account: {str(e)}')

        return redirect('admin_home')

    accounts = CustomUser.objects.all()
    return render(request, 'admin_home.html', {'accounts': accounts})


def toggle_status(request, report_id):
    if request.method == 'POST':
        report = get_object_or_404(InitialReport, id=report_id)
        
        if report.status == 'Ongoing':
            report.status = 'Case Closed'
        else:
            report.status = 'Ongoing'
        report.save()

        return JsonResponse({'status': report.status, 'message': 'Status updated successfully.'})

    return JsonResponse({'error': 'Invalid request method.'}, status=400)



def serve_css(request):
    return FileResponse(open('staticfiles/styles/style.css', 'rb'))


def login_view(request):
    if request.method == "POST":
        email = request.POST.get('email')
        password = request.POST.get('password')

        user = authenticate(request, username=email, password=password)

        if user is not None:
            if not user.verified:
                user.verified = True
                user.save()
                messages.success(request, "Your account has been verified.")

            login(request, user)

            if user.is_staff:
                return redirect('admin_home')
            elif user.is_active:
                return redirect('analytics')
            else:
                messages.error(request, "Your account is not active.")
        else:
            messages.error(request, "Invalid email or incorrect password. Try again")
    return render(request, 'login.html')



@login_required
def logout_view(request):
    logout(request)
    response = redirect('login')
    response.set_cookie('logged_out', 'true')
    response['Cache-Control'] = 'no-store'
    response['Pragma'] = 'no-cache'
    return response

@login_required
def analytics_view(request):
    context = {}  # Make sure context is defined
    return render(request, "analytics.html", context)

@login_required
def reports_view(request):
    reports = InitialReport.objects.all() 
    context = {
        'reports': reports,
    }
    return render(request, 'reports.html', context)


@login_required
def create_report_view(request):
    if request.method == 'POST':
        location = request.POST.get('crlocation')
        team = request.POST.get('crteam')
        time_reported = request.POST.get('crdetect')
        date_reported = request.POST.get('crdate')
        involved = request.POST.get('crinvolved')
        owner = request.POST.get('crowner')
        alarm_status = request.POST.get('cralarm')
        alarm_declared_by = request.POST.get('cralarm-dec')
        time_of_arrival = request.POST.get('crtime-arrive')
        time_under_control = request.POST.get('crtime-under')
        date_under_control = request.POST.get('crdate-under')
        under_control_declared_by = request.POST.get('crfunder-dec')
        time_out = request.POST.get('crtime-out')
        date_out = request.POST.get('crdate-out')
        out_declared_by = request.POST.get('crfout-dec')
        damages = request.POST.get('crdamages')
        fatality = request.POST.get('crfatality')  # Fatality input
        injured = request.POST.get('crinjured')
        families_affected = request.POST.get('craffected')
        establishments = request.POST.get('crestablishment')
        fire_trucks = request.POST.get('crtruck')
        ground_commander = request.POST.get('crground')
        ground_contact = request.POST.get('crground-num')
        safety_officer = request.POST.get('crsafety')
        safety_contact = request.POST.get('crsafety-num')
        sender = request.POST.get('crsender')
        sender_contact = request.POST.get('crsender-num')
        proof = request.FILES.get('modal-proof')

        def validate_date(date_str):
            if date_str:
                try:
                    return parse_date(date_str)
                except ValueError:
                    return None
            return None
        
        def validate_numeric(value):
            if value == '':
                return 0
            try:
                return int(value)
            except ValueError:
                raise ValidationError(f"Invalid number format for field '{value}'. It must be a valid number.")

        try:
            date_reported_valid = validate_date(date_reported)
            date_under_control_valid = validate_date(date_under_control)
            date_out_valid = validate_date(date_out)

            if not date_reported_valid:
                raise ValidationError("Invalid date format for 'Date Reported'. It must be in YYYY-MM-DD format.")
            if date_under_control and not date_under_control_valid:
                raise ValidationError("Invalid date format for 'Date Under Control'. It must be in YYYY-MM-DD format.")
            if date_out and not date_out_valid:
                raise ValidationError("Invalid date format for 'Date Out'. It must be in YYYY-MM-DD format.")

            fatality_valid = validate_numeric(fatality)
            injured_valid = validate_numeric(injured)
            families_affected_valid = validate_numeric(families_affected)
            establishments_valid = validate_numeric(establishments)
            fire_trucks_valid = validate_numeric(fire_trucks)

            try:
                time_reported_combined = datetime.strptime(f"{date_reported} {time_reported}", "%Y-%m-%d %H:%M")
            except ValueError:
                time_reported_combined = None  

            try:
                time_of_arrival_combined = datetime.strptime(f"{date_reported} {time_of_arrival}", "%Y-%m-%d %H:%M")
            except ValueError:
                time_of_arrival_combined = None

            try:
                time_under_control_combined = datetime.strptime(f"{date_under_control} {time_under_control}", "%Y-%m-%d %H:%M")
            except ValueError:
                time_under_control_combined = None

            try:
                time_out_combined = datetime.strptime(f"{date_out} {time_out}", "%Y-%m-%d %H:%M")
            except ValueError:
                time_out_combined = None

            new_report = InitialReport.objects.create(
                where=location,
                team=team,
                time_reported=time_reported_combined,
                date_reported=date_reported_valid,
                involved=involved,
                name_of_owner=owner,
                alarm_status=alarm_status,
                alarm_declared_by=alarm_declared_by,
                time_of_arrival=time_of_arrival_combined,
                time_of_fire_under_control=time_under_control_combined,
                date_of_fire_under_control=date_under_control_valid,
                fire_under_control_declared_by=under_control_declared_by,
                time_of_fire_out=time_out_combined,
                date_of_fire_out=date_out_valid,
                fire_out_declared_by=out_declared_by,
                estimated_damages=damages,
                no_of_fatality=fatality_valid,
                no_of_injured=injured_valid,
                no_of_families_affected=families_affected_valid,
                no_of_establishments=establishments_valid,
                no_of_fire_trucks=fire_trucks_valid,
                ground_commander=ground_commander,
                commander_contact_number=ground_contact,
                safety_officer=safety_officer,
                officer_contact_number=safety_contact,
                name_of_sender=sender,
                sender_contact_number=sender_contact,
                proof=proof,
                created_by=request.user
            )
            new_report.save()
            messages.success(request, 'Report has been submitted successfully.')
            return redirect('reports')

        except ValidationError as e:
            messages.error(request, f"Error submitting report: {e.message}")
            return render(request, 'create_reports.html')

        except Exception as e:
            messages.error(request, f"Error submitting report: {str(e)}")
            return render(request, 'create_reports.html')

    return render(request, "create_reports.html")

        
@login_required
def faq_view(request):
   return render(request, "faq.html")

def update_report(request, report_id):
    if request.method == 'POST':
        # Gather data from POST request, allowing fields to be optional
        location = request.POST.get('where')
        team = request.POST.get('team')
        date_reported = request.POST.get('date') or None
        time_reported = request.POST.get('detect') or None
        involved = request.POST.get('involved')
        owner = request.POST.get('owner')
        alarm = request.POST.get('alarm')
        alarm_declared_by = request.POST.get('alarm-dec')
        time_of_arrival = request.POST.get('time-arrive') or None
        time_of_fire_under_control = request.POST.get('time-under') or None
        date_of_fire_under_control = request.POST.get('date-under') or None
        fire_under_control_declared_by = request.POST.get('funder-dec')
        time_of_fire_out = request.POST.get('time-out') or None
        date_of_fire_out = request.POST.get('date-out') or None
        fire_out_declared_by = request.POST.get('fout-dec')
        estimated_damages = request.POST.get('damage', '0')
        no_of_fatality = request.POST.get('fatality', 0)
        no_of_injured = request.POST.get('injured', 0)
        no_of_families_affected = request.POST.get('affected', 0)
        no_of_establishments = request.POST.get('establishment', 1)
        no_of_fire_trucks = request.POST.get('truck', 0)
        ground_commander = request.POST.get('ground')
        commander_contact_number = request.POST.get('ground-num')
        safety_officer = request.POST.get('safety')
        officer_contact_number = request.POST.get('safety-num')
        name_of_sender = request.POST.get('sender')
        sender_contact_number = request.POST.get('sender-num')
        proof = request.FILES.get('proof')  

        # Fetch the report to be updated
        report = get_object_or_404(InitialReport, id=report_id)

        # Format datetime fields only if both date and time are provided
        datetime_of_report = f"{date_reported} {time_reported}" if date_reported and time_reported else None
        datetime_of_arrival = f"{date_reported} {time_of_arrival}" if date_reported and time_of_arrival else None
        datetime_of_fire_under_control = f"{date_reported} {time_of_fire_under_control}" if date_reported and time_of_fire_under_control else None
        datetime_of_fire_out = f"{date_of_fire_out} {time_of_fire_out}" if date_of_fire_out and time_of_fire_out else None

        # Update the report fields
        report.where = location
        report.team = team
        report.date_reported = date_reported
        report.time_reported = datetime_of_report
        report.involved = involved
        report.name_of_owner = owner
        report.alarm_status = alarm
        report.alarm_declared_by = alarm_declared_by
        report.time_of_arrival = datetime_of_arrival
        report.time_of_fire_under_control = datetime_of_fire_under_control
        report.date_of_fire_under_control = date_of_fire_under_control
        report.fire_under_control_declared_by = fire_under_control_declared_by
        report.time_of_fire_out = datetime_of_fire_out
        report.date_of_fire_out = date_of_fire_out
        report.fire_out_declared_by = fire_out_declared_by
        report.estimated_damages = estimated_damages
        report.no_of_fatality = no_of_fatality
        report.no_of_injured = no_of_injured
        report.no_of_families_affected = no_of_families_affected
        report.no_of_establishments = no_of_establishments
        report.no_of_fire_trucks = no_of_fire_trucks
        report.ground_commander = ground_commander
        report.commander_contact_number = commander_contact_number
        report.safety_officer = safety_officer
        report.officer_contact_number = officer_contact_number
        report.name_of_sender = name_of_sender
        report.sender_contact_number = sender_contact_number

        if proof:
            report.proof = proof  # Update proof if a new file is uploaded

        # Save the updated report to the database
        report.save()

        # Add success message and introduce a 2-second delay
        messages.success(request, 'Report has been updated successfully.')
        time.sleep(2)

        return redirect('reports')  # Redirect to reports page

    # If not a POST request, render the report editing page
    return render(request, 'reports.html')
        
@csrf_exempt  # Only if your desktop app doesn't support CSRF
def desktop_notification(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            # Store the update in cache
            cache.set('latest_update', data, timeout=None)
            # Store timestamp to track new updates
            cache.set('last_update_time', time.time(), timeout=None)
            return JsonResponse({"status": "success"})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

def event_stream(request):
    """SSE endpoint that clients connect to for updates"""
    def event_generator():
        last_check = time.time()
        
        while True:
            # Check if there's a new update
            current_time = cache.get('last_update_time')
            if current_time and current_time > last_check:
                data = cache.get('latest_update')
                if data:
                    # Send the update to the client
                    yield f"data: {json.dumps(data)}\n\n"
                    last_check = current_time
            
            time.sleep(1)  # Poll every second

    response = StreamingHttpResponse(event_generator(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['Connection'] = 'keep-alive'
    return response
