from django.shortcuts import render, HttpResponse, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from .forms import RegisterForms
from django.http import FileResponse, StreamingHttpResponse
from django.contrib import admin
from django.contrib import messages
from .models import CustomUser
from django.core.cache import cache
from django.core.validators import validate_email 
from .models import InitialReport, tempReports
from datetime import date,  timedelta
from django.db.models.functions import TruncMonth
from django.utils.timezone import now
from django.db.models import Count
from django.shortcuts import get_object_or_404, redirect
import random
import string
from django.core.mail import send_mail
from django.conf import settings
from django.http import JsonResponse
from django.contrib.auth.hashers import make_password
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.utils import timezone
from datetime import datetime
from datetime import  date
import json
import time
from django.core.exceptions import ValidationError
from django.utils.dateparse import parse_date
from django.core.validators import validate_integer
import time
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from django.db.models import Count
from django.db.models.functions import TruncMonth, TruncYear
from django.utils.timezone import now
from django.utils.timezone import localtime, now
from django.db.models import Count, Func
from django.db.models.functions import ExtractMonth
from django.db.models import Sum
from django.db.models import Count, Sum, Value
from django.db.models.functions import ExtractMonth, Coalesce
import logging
from .models import tempReports
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_protect
from django.utils.timezone import make_aware, get_current_timezone
from django.db.models.functions import ExtractYear
from supabase import create_client, Client
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.html import strip_tags
import mimetypes
import os
from .models import CustomUser

@csrf_exempt
def update_account(request):
    if request.method == "POST":
        # Get the account ID from the form
        account_id = request.POST.get("account_id")
        name = request.POST.get("name")
        email = request.POST.get("email")
        role = request.POST.get("role")
        contact_number = request.POST.get("contact_number")

        # Ensure all required fields are present
        if not account_id or not name or not email or not role:
            return JsonResponse({"error": "Missing required fields."}, status=400)

        # Get the account object
        account = get_object_or_404(CustomUser, id=account_id)

        # Update the account object
        account.name = name
        account.email = email
        account.role = role
        account.contact_number = contact_number

        account.save()

        return JsonResponse({"message": "Account updated successfully!"}, status=200)

    # If request method is not POST, return an error
    return JsonResponse({"error": "Invalid request method."}, status=400)


supabase: Client = create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)

def generate_filename(original_filename):
    # Extract the file extension from the original filename
    ext = os.path.splitext(original_filename)[1]
    
    # If no extension found, default to .jpg
    if not ext:
        ext = '.jpg'
    
    return f"fire_incident_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"

def fetch_names(request):
    if request.method == 'GET':
        # Fetch user ID, name, and contact number
        users = CustomUser.objects.filter(name__isnull=False).values('id', 'name', 'contact_number').distinct()
        return JsonResponse({'names': list(users)}, safe=False)


@csrf_protect
@require_POST
def transfer_report(request, temp_report_id):
    try:
        temp_report = tempReports.objects.get(id=temp_report_id)

        new_initial_report = InitialReport.objects.create(
            where=temp_report.where,
            date_reported=temp_report.date,
            time_reported=temp_report.time_detected,
            proof=temp_report.proof,
        )

        return JsonResponse({'status': 'success', 'report_id': new_initial_report.id})

    except tempReports.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Temp report not found'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
@csrf_protect
@require_POST
def update_temp_report_status(request, temp_report_id):
    try:
        data = json.loads(request.body)
        status = data.get('status')
        
        temp_report = tempReports.objects.get(id=temp_report_id)
        temp_report.status = status
        temp_report.save()
        
        return JsonResponse({'status': 'success'})
    except tempReports.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Report not found'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
def peak_report_summary(request):
    year = int(request.GET.get('year', datetime.now().year))
    
    reports = InitialReport.objects.filter(date_reported__year=year)
    
    month_counts = (
        reports
        .annotate(month=ExtractMonth('date_reported'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    peak_month = month_counts[0] if month_counts else None
    peak_month_name = datetime(1900, peak_month['month'], 1).strftime('%B') if peak_month else 'None'

    peak_month_reports = reports.filter(date_reported__month=peak_month['month']) if peak_month else None


    location_counts = (
        peak_month_reports
        .values('where')
        .annotate(count=Count('where'))  
        .order_by('-count')
    ) if peak_month_reports else None
    most_affected_location = location_counts[0]['where'] if location_counts else 'None'

    total_casualties = (
        peak_month_reports.aggregate(
            fatalities=Sum('no_of_fatality', default=0),
            injured=Sum('no_of_injured', default=0)
        ) if peak_month_reports else {'fatalities': 0, 'injured': 0}
    )
    total_casualties_count = (total_casualties['fatalities'] or 0) + (total_casualties['injured'] or 0)

    total_estimated_damages = 0
    if peak_month_reports:
        try:
            damages = peak_month_reports.values_list('estimated_damages', flat=True)
            total_estimated_damages = sum(
                int(damage) for damage in damages if damage and damage.isdigit()
            )
        except ValueError:
            total_estimated_damages = 'Invalid data'

    peak_data = {
        'peak_month': peak_month_name,
        'peak_month_count': peak_month['count'] if peak_month else 0,
        'reported_count' : reports.count(),
        'most_affected_location': most_affected_location,
        'total_casualties': total_casualties_count,
        'total_estimated_damages': total_estimated_damages,
    }

    return JsonResponse(peak_data)

def monthly_report_summary(request):
    year = int(request.GET.get('year', datetime.now().year))
    month = int(request.GET.get('month', datetime.now().month))

    reports = InitialReport.objects.filter(date_reported__year=year, date_reported__month=month)

    # Calculate the most affected location
    location_counts = (
        reports
        .values('where')
        .annotate(count=Count('where'))
        .order_by('-count')
    )
    most_affected_location = location_counts[0]['where'] if location_counts else 'None'

    # Calculate total casualties
    total_casualties = reports.aggregate(
        fatalities=Sum('no_of_fatality', default=0),
        injured=Sum('no_of_injured', default=0),
    )
    total_casualties_count = (total_casualties['fatalities'] or 0) + (total_casualties['injured'] or 0)

    # Calculate total estimated damages
    total_estimated_damages = 0
    try:
        damages = reports.values_list('estimated_damages', flat=True)
        total_estimated_damages = sum(
            int(damage) for damage in damages if damage and damage.isdigit()
        )
    except ValueError:
        total_estimated_damages = 'Invalid data'

    monthly_data = {
        'reported_count': reports.count(),
        'most_affected_location': most_affected_location,
        'total_casualties': total_casualties_count,
        'total_estimated_damages': total_estimated_damages,
    }

    return JsonResponse(monthly_data)

    
def reports_count_for_years(request):
    current_year = datetime.now().year
    start_year = current_year - 6  
    years_range = range(start_year, current_year + 1)

    counts = (
        InitialReport.objects
        .filter(date_reported__year__in=years_range)
        .annotate(year=ExtractYear('date_reported'))  
        .values('year')
        .annotate(count=Count('id'))  
        .order_by('year')
    )

    data = {year['year']: year['count'] for year in counts}

    for year in years_range:
        data.setdefault(year, 0)

    return JsonResponse({'yearly_data': data})

def monthly_reports_for_current_year(request):
    current_year = datetime.now().year

    reports_by_month = (
        InitialReport.objects.filter(date_reported__year=current_year)
        .annotate(month=TruncMonth('date_reported'))  
        .values('month')
        .annotate(count=Count('id'))  
        .order_by('month')
    )
    
    all_months = {date(current_year, month, 1).strftime('%b'): 0 for month in range(1, 13)}

    for report in reports_by_month:
        month_name = report['month'].strftime('%b')  
        all_months[month_name] = report['count']

    return JsonResponse({'year': current_year, 'monthly_data': all_months})

@login_required
def analytics_view(request):
    current_date = localtime(now())

    latest_reports = tempReports.objects.filter(
        status__in=['Reported', 'Dismissed']
    ).order_by('-date', '-time_detected')[:8]


    responded_this_month = InitialReport.objects.filter(
        status="Case Closed",
        date_reported__year=current_date.year,
        date_reported__month=current_date.month
    ).count()

    responded_this_year = InitialReport.objects.filter(
        status="Case Closed",
        date_reported__year=current_date.year
    ).count()

    print(f"Responded this month: {responded_this_month}")
    print(f"Responded this year: {responded_this_year}")
    print(f"Latest detections: {latest_reports}")

    context = {
        'responded_this_month': responded_this_month,
        'responded_this_year': responded_this_year,
        'latest_reports': latest_reports,
    }

    return render(request, "analytics.html", context)


def remove_accounts(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        emails = data.get('emails', [])
        
        if not emails:
            return JsonResponse({'success': False, 'message': 'No accounts selected.'}, status=400)

        deleted_count, _ = CustomUser.objects.filter(email__in=emails).delete()

        if deleted_count > 0:
            return JsonResponse({'success': True, 'message': 'Selected accounts have been removed successfully.'})
        else:
            return JsonResponse({'success': False, 'message': 'No accounts were found to remove.'}, status=404)

    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=400)


def export_initial_report(request):
    selected_year = request.GET.get('value', None)  

    print("Selected Year:", selected_year)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Initial Report Data"

    headers = [
        "No.", "Station", "Exact Location/ Address of Fire Incident", "Team", "Time Reported", "Date Reported",
        "Involved", "Name of Owner", "Alarm Status", "Alarm Declared By", "Time of Arrival At Scene",
        "Time of Fire Under Control", "Date of Fire Under Control", "Fire Under Control Declared By",
        "Time of Fire Out", "Date of Fire Out", "Fire Out Declared By", "Estimated Damages", "No. of Fatalities",
        "No. of Injured", "No. of Families Affected", "No. of Establishments", "No. of Fire Trucks",
        "Ground Commander", "Commander Contact Number", "Safety Officer", "Officer Contact Number",
        "Name of Sender", "Sender Contact Number", "Remarks"
    ]

    fields = [
        'where', 'team', 'time_reported', 'date_reported', 'involved', 'name_of_owner',
        'alarm_status', 'alarm_declared_by', 'time_of_arrival', 'time_of_fire_under_control',
        'date_of_fire_under_control', 'fire_under_control_declared_by', 'time_of_fire_out',
        'date_of_fire_out', 'fire_out_declared_by', 'estimated_damages', 'no_of_fatality',
        'no_of_injured', 'no_of_families_affected', 'no_of_establishments', 'no_of_fire_trucks',
        'ground_commander', 'commander_contact_number', 'safety_officer', 'officer_contact_number',
        'name_of_sender', 'sender_contact_number', 'status'
    ]

    header_font = Font(bold=True, color="000000")
    data_fill = PatternFill(start_color="DF4B40", end_color="DF4B40", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thick_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    for col_num, header_item in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_item)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thick_border

    ws.row_dimensions[1].height = 30

    sy_text = f"S/Y {selected_year}" if selected_year else "S/Y All Years"
    sy_cell = ws.cell(row=2, column=1, value=sy_text)
    sy_cell.font = Font(bold=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sy_cell.alignment = center_alignment  
    sy_cell.border = thick_border

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = yellow_fill
        cell.alignment = center_alignment
        cell.border = thick_border

    reports = InitialReport.objects.annotate(report_year=ExtractYear('date_reported'))
    if selected_year:
        reports = reports.filter(report_year=int(selected_year))
    reports = reports.filter(status__in=['Case Closed', 'Ongoing']).values(*fields)

    for row_num, report in enumerate(reports, 4):  
        no_value = row_num - 3  
        station_value = "Santa Rosa City"

        no_cell = ws.cell(row=row_num, column=1, value=no_value)
        no_cell.fill = data_fill
        no_cell.alignment = center_alignment
        no_cell.border = thick_border

        station_cell = ws.cell(row=row_num, column=2, value=station_value)
        station_cell.fill = data_fill
        station_cell.alignment = center_alignment
        station_cell.border = thick_border

        for col_num, field in enumerate(fields, 3):
            value = report.get(field, "None")

            if not value and field in ['no_of_fatality', 'no_of_injured', 'no_of_families_affected', 'no_of_establishments', 'no_of_fire_trucks']:
                value = 0
            elif not value:
                value = "None"

            if isinstance(value, datetime):
                value = value.replace(tzinfo=None)

            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = data_fill
            cell.alignment = center_alignment
            cell.border = thick_border

    for col in range(2, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    ws.column_dimensions['A'].width = 5  

    file_name = f"initial_report_{selected_year}.xlsx" if selected_year else "initial_report_all_years.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    wb.save(response)
    return response


@login_required
def change_password(request):
    if request.method == 'POST':
        new_password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')

        if new_password != confirm_password:
            messages.error(request, 'Passwords do not match.')
            return redirect(request.META.get('HTTP_REFERER', 'analytics'))

        user = request.user

        user.set_password(new_password)
        user.save()

        update_session_auth_hash(request, user)

        # Prepare email context
        context = {
            'user': user,
            'date': timezone.now(),
            'year': timezone.now().year,
        }

        # Render HTML email
        html_message = render_to_string('emails/forgot_password_confirmation.html', context)
        plain_message = strip_tags(html_message)  # Create plain text version

        # Send email
        send_mail(
            'Password Reset Successful',
            plain_message,
            settings.DEFAULT_FROM_EMAIL,
            [user.email],
            html_message=html_message,
            fail_silently=False,
        )
        
        messages.success(request, 'Password reset successful! A confirmation email has been sent.')

        return redirect(request.META.get('HTTP_REFERER', 'analytics'))

    return render(request, 'analytics.html')


def generate_random_password(length=10):
    """Generates a random password."""
    characters = string.ascii_letters + string.digits + string.punctuation
    return ''.join(random.choice(characters) for i in range(length))


def forgot_password_view(request):
    show_success_modal = False

    if request.method == 'POST':
        email = request.POST.get('email')
        new_password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')

        if new_password != confirm_password:
            messages.error(request, 'Passwords do not match.')
            return redirect('forgot_password')

        try:
            user = CustomUser.objects.get(email=email)
            user.set_password(new_password)
            user.save()

            # Prepare email context
            context = {
                'user': user,
                'date': timezone.now(),
                'year': timezone.now().year,
            }

            # Render HTML email
            html_message = render_to_string('emails/forgot_password_confirmation.html', context)
            plain_message = strip_tags(html_message)  # Create plain text version

            # Send email
            send_mail(
                'Password Reset Confirmation',
                plain_message,
                settings.DEFAULT_FROM_EMAIL,
                [email],
                html_message=html_message,
                fail_silently=False,
            )

            messages.success(request, 'Password reset successful! A confirmation email has been sent.')
            show_success_modal = True

        except CustomUser.DoesNotExist:
            messages.error(request, 'Email address not found.')
            return redirect('forgot_password')

    return render(request, 'forgot_password.html', {'show_success_modal': show_success_modal})


def check_email_exists(request):
    email = request.GET.get('email', None)
    phone = request.GET.get('phone_number', None)  # Get the phone number from the request

    response_data = {}

    if email:
        response_data['email_exists'] = CustomUser.objects.filter(email=email).exists()

    if phone:
        response_data['phone_exists'] = CustomUser.objects.filter(contact_number=phone).exists()

    return JsonResponse(response_data)


@login_required
def register_view(request):
    account_added = False
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        role = request.POST.get('role')
        contact_number = request.POST.get('contact_number')

        if not all([name, email, role, contact_number]):
            messages.error(request, 'All fields are required.')
        else:
            try:
                if not contact_number.isdigit() or len(contact_number) not in [10, 11]:
                    raise ValidationError("Contact number must be 10-11 digits.")

                generated_password = generate_random_password()

                new_user = CustomUser.objects.create_user(
                    name=name,
                    email=email,
                    role=role,
                    contact_number=contact_number,
                    password=generated_password
                )

                # Prepare email context
                context = {
                    'name': name,
                    'email': email,
                    'role': role,
                    'contact_number': contact_number,
                    'password': generated_password,
                    'date': timezone.now(),
                    'year': timezone.now().year,
                }

                # Render email template
                html_message = render_to_string('emails/registration_confirmation.html', context)
                plain_message = strip_tags(html_message)

                # Send email
                send_mail(
                    'Welcome to TruetheFire - Your Account Details',
                    plain_message,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    html_message=html_message,
                    fail_silently=False,
                )

                account_added = True
                messages.success(request, 'Account has been added successfully.')
                return render(request, 'admin_home.html', {'accounts': CustomUser.objects.all(), 'account_added': account_added})
            except ValidationError as e:
                messages.error(request, f'Validation error: {e.message}')
            except Exception as e:
                messages.error(request, f'Error adding account: {str(e)}')

    accounts = CustomUser.objects.all().order_by('name')
    return render(request, 'admin_home.html', {'accounts': accounts, 'account_added': account_added})


def toggle_status(request, report_id):
    if request.method == 'POST':
        report = get_object_or_404(InitialReport, id=report_id)
        print("CSRF Token:", request.META.get("HTTP_X_CSRFTOKEN"))
        
        if report.status == 'Ongoing':
            report.status = 'Case Closed'
        else:
            report.status = 'Ongoing'
        report.save()

        return JsonResponse({'status': report.status, 'message': 'Status updated successfully.'})

    return JsonResponse({'error': 'Invalid request method.'}, status=400)



def serve_css(request):
    return FileResponse(open('staticfiles/styles/style.css', 'rb'))


def login_view(request):
    if request.method == "POST":
        email = request.POST.get('email')
        password = request.POST.get('password')

        user = authenticate(request, username=email, password=password)

        if user is not None:
            if not user.verified:
                user.verified = True
                user.save()
                messages.success(request, "Your account has been verified.")

            login(request, user)

            if user.is_staff:
                return redirect('admin_home')
            elif user.is_active:
                return redirect('analytics')
            else:
                messages.error(request, "Your account is not active.")
        else:
            messages.error(request, "Invalid email or incorrect password. Try again")
    return render(request, 'login.html')



@login_required
def logout_view(request):
    logout(request)
    response = redirect('login')
    response.set_cookie('logged_out', 'true')
    response['Cache-Control'] = 'no-store'
    response['Pragma'] = 'no-cache'
    return response

@csrf_exempt
def archive_report(request):
    if request.method == 'POST':
        report_ids = request.POST.get('report_ids', '').split(',')

        if not report_ids:
            return JsonResponse({'success': False, 'message': 'No report IDs provided.'}, status=400)

        try:
            reports = InitialReport.objects.filter(id__in=report_ids)
            reports.update(is_archived=True)

            return JsonResponse({'success': True, 'message': 'Reports archived successfully.'})

        except InitialReport.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'One or more reports not found.'}, status=404)

    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=400)


@login_required
def reports_view(request):
    reports = InitialReport.objects.filter(is_archived=False).order_by('-date_reported', '-time_reported')
    archived_Reports = InitialReport.objects.filter(is_archived=True).order_by('-date_reported', '-time_reported')

    context = {
        'reports': reports,
        'archived_Reports': archived_Reports
    }
    return render(request, 'reports.html', context)




@login_required
def create_report_view(request):
    if request.method == 'POST':
        location = request.POST.get('crlocation')
        team = request.POST.get('crteam')
        time_reported = request.POST.get('crdetect')
        date_reported = request.POST.get('crdate')
        involved = request.POST.get('crinvolved')
        owner = request.POST.get('crowner')
        alarm_status = request.POST.get('cralarm')
        alarm_declared_by = request.POST.get('cralarm-dec')
        time_of_arrival = request.POST.get('crtime-arrive')
        time_under_control = request.POST.get('crtime-under')
        date_under_control = request.POST.get('crdate-under')
        under_control_declared_by = request.POST.get('crfunder-dec')
        time_out = request.POST.get('crtime-out')
        date_out = request.POST.get('crdate-out')
        out_declared_by = request.POST.get('crfout-dec')
        damages = request.POST.get('crdamages')
        fatality = request.POST.get('crfatality')  # Fatality input
        injured = request.POST.get('crinjured')
        families_affected = request.POST.get('craffected')
        establishments = request.POST.get('crestablishment')
        fire_trucks = request.POST.get('crtruck')
        ground_commander = request.POST.get('crground')
        ground_contact = request.POST.get('crground-num')
        safety_officer = request.POST.get('crsafety')
        safety_contact = request.POST.get('crsafety-num')
        sender = request.POST.get('crsender')
        sender_contact = request.POST.get('crsender-num')
        proof = request.FILES.get('modal-proof')

        # If a file is uploaded, upload to Supabase and get the URL
        file_url = None
        if proof:
            file_name = generate_filename(proof.name)  # Get the file name
            file_path = file_name  # Path in the bucket
            
            file_data = proof.read()  # Read the file contents

            response = supabase.storage.from_('FireProof').upload(file_path, file_data)

            file_url = supabase.storage.from_('FireProof').get_public_url(file_path)

        def validate_date(date_str):
            if date_str:
                try:
                    return parse_date(date_str)
                except ValueError:
                    return None
            return None
        
        def validate_numeric(value):
            if value == '':
                return 0
            try:
                return int(value)
            except ValueError:
                raise ValidationError(f"Invalid number format for field '{value}'. It must be a valid number.")

        try:
            # Validate dates
            date_reported_valid = validate_date(date_reported)
            date_under_control_valid = validate_date(date_under_control)
            date_out_valid = validate_date(date_out)

            if not date_reported_valid:
                raise ValidationError("Invalid date format for 'Date Reported'. It must be in YYYY-MM-DD format.")
            if date_under_control and not date_under_control_valid:
                raise ValidationError("Invalid date format for 'Date Under Control'. It must be in YYYY-MM-DD format.")
            if date_out and not date_out_valid:
                raise ValidationError("Invalid date format for 'Date Out'. It must be in YYYY-MM-DD format.")

            # Validate numbers
            fatality_valid = validate_numeric(fatality)
            injured_valid = validate_numeric(injured)
            families_affected_valid = validate_numeric(families_affected)
            establishments_valid = validate_numeric(establishments)
            fire_trucks_valid = validate_numeric(fire_trucks)

            # Combine date and time fields
            try:
                time_reported_combined = datetime.strptime(f"{date_reported} {time_reported}", "%Y-%m-%d %H:%M")
            except ValueError:
                time_reported_combined = None  

            try:
                time_of_arrival_combined = datetime.strptime(f"{date_reported} {time_of_arrival}", "%Y-%m-%d %H:%M")
            except ValueError:
                time_of_arrival_combined = None

            try:
                time_under_control_combined = datetime.strptime(f"{date_under_control} {time_under_control}", "%Y-%m-%d %H:%M")
            except ValueError:
                time_under_control_combined = None

            try:
                time_out_combined = datetime.strptime(f"{date_out} {time_out}", "%Y-%m-%d %H:%M")
            except ValueError:
                time_out_combined = None

            last_report = InitialReport.objects.order_by('id').last()

            if last_report and getattr(last_report, 'fir_number', None):  # Check if last_report and fir_number exist
                try:
                    last_id = int(last_report.fir_number.split('-')[1])  # Safely extract the number part
                    new_id = last_id + 1
                except (ValueError, IndexError):  # Handle invalid or malformed fir_number
                    new_id = 1
            else:
                new_id = 1

            fir_number = f"FIR-{new_id:02d}"

            # Create report with Supabase file URL
            new_report = InitialReport.objects.create(
                fir_number=fir_number, 
                where=location,
                team=team,
                time_reported=time_reported_combined,
                date_reported=date_reported_valid,
                involved=involved,
                name_of_owner=owner,
                alarm_status=alarm_status,
                alarm_declared_by=alarm_declared_by,
                time_of_arrival=time_of_arrival_combined,
                time_of_fire_under_control=time_under_control_combined,
                date_of_fire_under_control=date_under_control_valid,
                fire_under_control_declared_by=under_control_declared_by,
                time_of_fire_out=time_out_combined,  
                date_of_fire_out=date_out_valid,
                fire_out_declared_by=out_declared_by,
                estimated_damages=damages,
                no_of_fatality=fatality_valid,
                no_of_injured=injured_valid,
                no_of_families_affected=families_affected_valid,
                no_of_establishments=establishments_valid,
                no_of_fire_trucks=fire_trucks_valid,
                ground_commander=ground_commander,
                commander_contact_number=ground_contact,
                safety_officer=safety_officer,
                officer_contact_number=safety_contact,
                name_of_sender=sender,
                sender_contact_number=sender_contact,
                proof=file_url,  # Store the Supabase public URL
                created_by=request.user
            )
            new_report.save()

            # Add context variable to trigger the modal in template
            messages.success(request, 'Report has been submitted successfully.')
            return render(request, 'create_reports.html', {'report_added': True})  # Pass context variable

        except ValidationError as e:
            messages.error(request, f"Error submitting report: {e.message}")
            return render(request, 'create_reports.html')

        except Exception as e:
            messages.error(request, f"Error submitting report: {str(e)}")
            return render(request, 'create_reports.html')

    return render(request, "create_reports.html", {'report_added': False})  # Default value
        
@login_required
def faq_view(request):
   return render(request, "faq.html")


logger = logging.getLogger(__name__)

def convert_to_aware_datetime(date_str, time_str):
    """Convert date and time strings to a timezone-aware datetime."""
    if not date_str or not time_str:  # Check for blank or None values
        return None
    try:
        naive_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
        return make_aware(naive_dt, get_current_timezone())
    except ValueError as e:
        print(f"Invalid datetime format: {date_str} {time_str}. Error: {e}")
        return None
    
def update_report(request, report_id):
    if request.method == 'POST':
        report = get_object_or_404(InitialReport, id=report_id)
        data = request.POST


        report.where = data.get('where', report.where)
        report.team = data.get('team', report.team)
        report.date_reported = data.get('date', report.date_reported)
        report.alarm_status = data.get('alarm', report.alarm_status)

        report.name_of_owner = data.get('owner', None) if data.get('owner') != "None" else None
        report.alarm_declared_by = data.get('alarm-dec', None) if data.get('alarm-dec') != "None" else None
        report.fire_under_control_declared_by = data.get('funder-dec', None) if data.get('funder-dec') != "None" else None
        report.fire_out_declared_by = data.get('fout-dec', None) if data.get('fout-dec') != "None" else None
        report.involved = data.get('involved', None) if data.get('involved') != "None" else None

        def convert_to_aware_datetime(date_str, time_str):
            if date_str and time_str:
                naive_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
                return make_aware(naive_dt)
            return None

        report.time_reported = convert_to_aware_datetime(data.get('date', ''), data.get('detect', '')) or report.time_reported
        report.time_of_arrival = convert_to_aware_datetime(data.get('date', '').strip(), data.get('time-arrive', '').strip()) or report.time_of_arrival
        report.time_of_fire_under_control = convert_to_aware_datetime(data.get('date', '').strip(), data.get('time-under', '').strip()) or report.time_of_fire_under_control
        report.time_of_fire_out = convert_to_aware_datetime(data.get('date', '').strip(), data.get('time-out', '').strip()) or report.time_of_fire_out

        def get_int_value(field_name, default=0):
            try:
                raw_value = data.get(field_name, '').replace(',', '').strip()  
                return int(raw_value or default)
            except (ValueError, TypeError):
                return default

        report.estimated_damages = get_int_value('damage', report.estimated_damages or 0)
        report.no_of_fatality = get_int_value('fatality', report.no_of_fatality or 0)
        report.no_of_injured = get_int_value('injured', report.no_of_injured or 0)
        report.no_of_families_affected = get_int_value('affected', report.no_of_families_affected or 0)
        report.no_of_establishments = get_int_value('establishment', report.no_of_establishments or 0)
        report.no_of_fire_trucks = get_int_value('truck', report.no_of_fire_trucks or 0)

        report.ground_commander = data.get('ground', '').strip() or report.ground_commander
        report.commander_contact_number = data.get('ground-num', '').strip() or report.commander_contact_number
        report.safety_officer = data.get('safety', '').strip() or report.safety_officer
        report.officer_contact_number = data.get('safety-num', '').strip() or report.officer_contact_number
        report.name_of_sender = data.get('sender', '').strip() or report.name_of_sender
        report.sender_contact_number = data.get('sender-num', '').strip() or report.sender_contact_number

        proof = request.FILES.get('input-proof')
        if proof:
            try:
                file_name = generate_filename(proof.name)
                file_path = f"proofs/{file_name}"

                # Read file data
                file_data = proof.read()

                # Upload to Supabase
                response = supabase.storage.from_('FireProof').upload(
                    file_path, 
                    file_data, 
                )

                # Get public URL
                file_url = supabase.storage.from_('FireProof').get_public_url(file_path)
                
                # Update report with new file URL
                report.proof = file_url

            except Exception as e:
                print(f"Error uploading proof: {e}")
                return JsonResponse({
                    'status': 'error', 
                    'message': f'Failed to upload proof: {str(e)}'
                }, status=500)
            
        try:
            report.save()
            return JsonResponse({'status': 'success', 'message': 'Report updated successfully.'})
        except Exception as e:
            print(f"Error saving report: {e}")
            return JsonResponse({'status': 'error', 'message': 'Failed to save report.'}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=400)

        
@csrf_exempt
def desktop_notification(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            cache.set('latest_update', data, timeout=None)
            cache.set('last_update_time', time.time(), timeout=None)
            return JsonResponse({"status": "success"})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

def event_stream(request):
    """SSE endpoint that clients connect to for updates"""
    def event_generator():
        last_check = time.time()
        
        while True:
            current_time = cache.get('last_update_time')
            if current_time and current_time > last_check:
                data = cache.get('latest_update')
                if data:
                    yield f"data: {json.dumps(data)}\n\n"
                    last_check = current_time
            
            time.sleep(1)

    response = StreamingHttpResponse(event_generator(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['Connection'] = 'keep-alive'
    return response

def get_temp_report_details(report_id):
    try:
        report = tempReports.objects.get(id=report_id)
        return {
            'id': report.id,
            'where': report.where,
            'date': report.date.strftime('%B %d, %Y'),
            'time': report.time_detected.strftime('%I:%M %p'),
            'proof': report.proof,
            'status': report.status
        }
    except tempReports.DoesNotExist:
        return None

def get_temp_report_details_view(request, report_id):
    details = get_temp_report_details(report_id)
    if details:
        return JsonResponse(details)
    return JsonResponse({'error': 'Report not found'}, status=404)

def get_unresolved_reports(request):
    # Query tempReports where status is null
    unresolved_reports = tempReports.objects.filter(status__isnull=True)
    
    # Prepare report details
    report_details = []
    for report in unresolved_reports:
        report_details.append({
            'id': report.id,
            'where': report.where,
            'date': report.date.strftime('%B %d, %Y'),
            'time': report.time_detected.strftime('%I:%M %p'),
            'proof': report.proof,
            'status': report.status
        })
    
    return JsonResponse(report_details, safe=False)